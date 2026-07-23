import openpyxl
from pathlib import Path
import model

cur_dir = Path(__file__).parent

def parse_can_matrix(ws):
    messages: list[Message] = []
    current_message: Message = None
    prev_network: str = None
    prev_message_name: str = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        network = row[0]
        message_name = row[1]

        if (network != prev_network) or (message_name != prev_message_name):
            if current_message is not None:
                messages.append(current_message)
            current_message = model.Message(
                message_name=row[1],
                message_id=int(row[2], 16),
                dlc=row[5],
                cyc_time=row[6],
                transmitter=row[8],
                signals=[],
                is_extended=(row[3] != "Standard"),
                is_canfd=(row[4] != "Classic CAN"),
            )

        new_signal = model.Signal(
            signal_name=row[9],
            start_bit=row[11],
            length=row[12],
            byte_order=model.ByteOrder.INTEL if row[13] == "Intel" else model.ByteOrder.MOTOROLA,
            signed=(row[14] == "Yes"),
            factor=row[15],
            offset=row[16],
            min_value=row[17],
            max_value=row[18],
            unit=row[19],
            receivers=row[21].split(',')
        )
        current_message.signals.append(new_signal)

        prev_network = network
        prev_message_name = message_name

    if current_message is not None:
        messages.append(current_message)

    return model.DataBase(messages)

def parse_load_workbook():
    wb = openpyxl.load_workbook(cur_dir / "input" / "CAN_Matrix_Sample_v2.xlsx")
    ws = wb["CAN Matrix"]
    database = parse_can_matrix(ws)
    return database
    # print(database)

# if __name__ == "__main__":
#     wb = openpyxl.load_workbook(cur_dir / "input" / "CAN_Matrix_Sample.xlsx")
#     ws = wb["CAN Matrix"]
#     database = parse_can_matrix(ws)
#     print(database)